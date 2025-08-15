import sys
import shutil
import datetime
import re
import traceback
import threading
import base64  # 用于嵌入图标
import os
from pathlib import Path
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QLabel,
    QLineEdit, QPushButton, QComboBox, QCheckBox, QGroupBox, QFileDialog,
    QMessageBox, QTreeWidget, QTreeWidgetItem, QProgressBar, QDialog,
    QDialogButtonBox, QListWidget, QListWidgetItem, QAbstractItemView,
    QScrollArea, QMenu, QAction, QInputDialog,
    QRadioButton, QButtonGroup, QHeaderView
)
from PyQt5.QtGui import QIcon, QColor, QPalette, QLinearGradient, QBrush, QFont, QPixmap, QPainter
from PyQt5.QtCore import Qt, QPoint, QByteArray
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# 注册支持多语言的字体
def register_multilingual_fonts():
    system_fonts = {
        "SimSun": "C:/Windows/Fonts/simsun.ttc",         # 宋体常规
        "SimSun-Bold": "C:/Windows/Fonts/simsunb.ttf",   # 宋体粗体
        "Microsoft YaHei": "C:/Windows/Fonts/msyh.ttc",  # 微软雅黑
        "Arial": "C:/Windows/Fonts/arial.ttf"            # Arial
    }

    for font_name, font_path in system_fonts.items():
        try:
            if Path(font_path).exists():
                pdfmetrics.registerFont(TTFont(font_name, font_path))
                print(f"成功注册字体: {font_name}")
            else:
                print(f"字体文件不存在: {font_path}")
        except Exception as e:
            print(f"注册字体 {font_name} 时出错: {str(e)}")

    if not pdfmetrics.getRegisteredFontNames():
        try:
            pdfmetrics.registerFont(TTFont('Vera', 'Vera.ttf'))
            print("使用 ReportLab 默认字体 Vera")
        except:
            pass

class FileGatherPro(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("文件归集管理器")
        # 使用app.ico文件作为图标
        icon_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "app.ico")
        if os.path.exists(icon_path):
            self.setWindowIcon(QIcon(icon_path))
        else:
            # 如果找不到ico文件，使用默认图标
            self.setWindowIcon(QIcon())
        self.setGeometry(100, 100, 1000, 750)
        
        # 修改1: 设置最小和最大宽度相同，高度可调
        self.setMinimumSize(1000, 600)  # 最小高度600
        self.setMaximumWidth(1000)       # 固定宽度
        
        # 初始化变量
        self.search_results = []
        self.target_folder = ""
        self.version = "2.3.4"  # 更新版本号为2.3.4
        self.update_log = "优化搜索结果及按钮显示 (2025-07-17)"
        self.search_folders = []
        self.found_files_count = 0
        self.searching = False
        self.cancel_search = False
        self.operation_log = []  # 操作日志
        self.operated_files = set()  # 操作过的文件名集合
        self.add_log("启动程序")
        
        register_multilingual_fonts()
        self.setup_gradient_background()
        self.init_ui()
        
        self.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                border: 1px solid #3498db;
                border-radius: 5px;
                margin-top: 1ex;
                background-color: rgba(255, 255, 255, 180);
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                subcontrol-position: top center;
                padding: 0 5px;
                color: #2980b9;
            }
            QPushButton {
                background-color: #3498db;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
            QPushButton:pressed {
                background-color: #1c6ea4;
            }
            QPushButton:disabled {
                background-color: #95a5a6;
            }
            QTreeWidget {
                background-color: rgba(255, 255, 255, 200);
                border: 1px solid #3498db;
                border-radius: 3px;
            }
            QLineEdit, QComboBox, QDateEdit {
                padding: 5px;
                border: 1px solid #3498db;
                border-radius: 3px;
                background-color: rgba(255, 255, 255, 200);
            }
            QProgressBar {
                border: 1px solid #3498db;
                border-radius: 3px;
                text-align: center;
                background-color: rgba(255, 255, 255, 200);
            }
            QProgressBar::chunk {
                background-color: #3498db;
                width: 10px;
            }
            QListWidget {
                background-color: rgba(255, 255, 255, 200);
                border: 1px solid #3498db;
                border-radius: 3px;
            }
            QRadioButton {
                padding: 4px;
            }
        """)
        
    def setup_gradient_background(self):
        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)
        
        palette = self.palette()
        gradient = QLinearGradient(0, 0, 0, self.height())
        gradient.setColorAt(0, QColor(230, 245, 255))
        gradient.setColorAt(1, QColor(180, 220, 255))
        palette.setBrush(QPalette.Window, QBrush(gradient))
        self.setPalette(palette)
        
    def init_ui(self):
        main_layout = QVBoxLayout()
        main_layout.setSpacing(15)
        main_layout.setContentsMargins(20, 20, 20, 20)
        
        title_label = QLabel("文件归集管理器")
        title_label.setFont(QFont("Arial", 18, QFont.Bold))
        title_label.setAlignment(Qt.AlignCenter)
        title_label.setStyleSheet("color: #2c3e50;")
        main_layout.addWidget(title_label)
        
        version_label = QLabel(f"版本: V{self.version} | © 2025 D&Ai/2FX 文件归集管理器")
        version_label.setAlignment(Qt.AlignCenter)
        version_label.setStyleSheet("color: #7f8c8d; font-size: 10pt;")
        main_layout.addWidget(version_label)
        
        search_group = QGroupBox("搜索条件")
        search_layout = QVBoxLayout()
        
        folder_layout = QVBoxLayout()
        folder_label_layout = QHBoxLayout()
        folder_label = QLabel("搜索文件夹:")
        
        # 修改2: 设置文件夹列表最大高度为3行
        self.folder_list = QListWidget()
        self.folder_list.setMaximumHeight(75)  # 大约3行高度
        
        folder_button_layout = QHBoxLayout()
        self.add_folder_button = QPushButton("添加文件夹")
        self.add_folder_button.clicked.connect(self.add_search_folder)
        self.add_drive_button = QPushButton("添加盘符")
        self.add_drive_button.clicked.connect(self.add_drive)
        
        self.remove_folder_button = QPushButton("删除选中")
        self.remove_folder_button.clicked.connect(self.remove_selected_folders)
        
        self.clear_folders_button = QPushButton("清空列表")
        self.clear_folders_button.clicked.connect(self.clear_search_folders)
        
        folder_button_layout.addWidget(self.add_folder_button)
        folder_button_layout.addWidget(self.add_drive_button)
        folder_button_layout.addWidget(self.remove_folder_button)
        folder_button_layout.addWidget(self.clear_folders_button)
        
        folder_label_layout.addWidget(folder_label)
        folder_layout.addLayout(folder_label_layout)
        folder_layout.addWidget(self.folder_list)
        folder_layout.addLayout(folder_button_layout)
        search_layout.addLayout(folder_layout)
        
        keyword_layout = QHBoxLayout()
        keyword_label = QLabel("关键词:")
        self.keyword_entry = QLineEdit()
        self.keyword_entry.setPlaceholderText("输入文件名包含的关键词")
        self.keyword_entry.setToolTip(
            "支持多种搜索模式：\n"
            "- 基本搜索: 输入关键词，如 \"报告\"\n"
            "- 多关键词: 用空格分隔多个关键词，如 \"报告 2025\"\n"
            "- 逻辑与: 使用 \"+\" 表示必须包含，如 \"+重要 +财务\"\n"
            "- 逻辑或: 使用 \"|\" 表示或关系，如 \"报告|总结\"\n"
            "- 排除: 使用 \"-\" 排除关键词，如 \"报告 -草稿\"\n"
            "- 通配符: 使用 \"*\" 匹配任意字符，如 \"项目*报告\"\n"
            "- 精确匹配: 使用引号进行精确匹配，如 \"\\\"季度报告\\\"\"\n"
            "\n示例: \"项目 +最终版 -草稿\" 表示搜索包含\"项目\"和\"最终版\"但不包含\"草稿\"的文件"
        )
        
        keyword_layout.addWidget(keyword_label)
        keyword_layout.addWidget(self.keyword_entry, 1)
        search_layout.addLayout(keyword_layout)
        
        search_mode_layout = QHBoxLayout()
        search_mode_label = QLabel("搜索模式:")
        
        self.search_mode_group = QButtonGroup(self)
        self.filename_radio = QRadioButton("仅文件名")
        self.content_radio = QRadioButton("仅内容")
        self.both_radio = QRadioButton("两者同时")
        self.filename_radio.setChecked(True)
        
        tooltip = (
            "选择搜索模式：\n"
            "- 仅文件名: 只在文件名中匹配关键词\n"
            "- 仅内容: 只在文件内容中匹配关键词\n"
            "- 两者同时: 在文件名或内容中匹配关键词即可\n\n"
            "注意：内容搜索仅检查文件的前3000个字符"
        )
        self.filename_radio.setToolTip(tooltip)
        self.content_radio.setToolTip(tooltip)
        self.both_radio.setToolTip(tooltip)
        
        search_mode_layout.addWidget(search_mode_label)
        search_mode_layout.addWidget(self.filename_radio)
        search_mode_layout.addWidget(self.content_radio)
        search_mode_layout.addWidget(self.both_radio)
        search_mode_layout.addStretch(1)
        
        search_layout.addLayout(search_mode_layout)
        
        filetype_layout = QHBoxLayout()
        filetype_label = QLabel("文件类型:")
        self.filetype_combo = QComboBox()
        self.filetype_combo.addItem("所有文件", "")
        self.filetype_combo.addItem("文档", [".doc", ".docx", ".txt", ".pdf", ".xls", ".xlsx", ".ppt", ".pptx"])
        self.filetype_combo.addItem("图片", [".jpg", ".jpeg", ".png", ".gif", ".bmp", ".tiff", ".webp"])
        self.filetype_combo.addItem("视频", [".mp4", ".avi", ".mov", ".mkv", ".flv", ".wmv", ".mpg"])
        self.filetype_combo.addItem("音频", [".mp3", ".wav", ".flac", ".aac", ".ogg", ".wma"])
        self.filetype_combo.addItem("可执行文件", [".exe", ".msi", ".bat", ".cmd"])
        self.filetype_combo.addItem("压缩文件", [".zip", ".rar", ".7z", ".tar", ".gz"])
        self.filetype_combo.addItem("自定义", "custom")
        self.filetype_combo.setToolTip(
            "选择要搜索的文件类型：\n"
            "- 所有文件: 搜索功能所有类型的文件\n"
            "- 文档: 包括 .doc, .docx, .txt, .pdf 等\n"
            "- 图片: 包括 .jpg, .jpeg, .png, .gif 等\n"
            "- 视频: 包括 .mp4, .avi, .mov, .mkv 等\n"
            "- 音频: 包括 .mp3, .wav, .flac, .aac 等\n"
            "- 可执行文件: 包括 .exe, .msi, .bat, .cmd 等\n"
            "- 压缩文件: 包括 .zip, .rar, .7z, .tar 等\n"
            "- 自定义: 手动输入扩展名，用分号分隔，如 \".py;.java;.cpp\""
        )
        
        filetype_layout.addWidget(filetype_label)
        filetype_layout.addWidget(self.filetype_combo, 1)
        search_layout.addLayout(filetype_layout)
        
        date_size_layout = QHBoxLayout()
        
        mod_date_layout = QVBoxLayout()
        mod_date_label = QLabel("修改日期:")
        self.mod_date_combo = QComboBox()
        self.mod_date_combo.addItem("不限", (None, None))
        self.mod_date_combo.addItem("今天", (datetime.date.today(), datetime.date.today()))
        self.mod_date_combo.addItem("最近7天", (datetime.date.today() - datetime.timedelta(days=7), datetime.date.today()))
        self.mod_date_combo.addItem("最近30天", (datetime.date.today() - datetime.timedelta(days=30), datetime.date.today()))
        self.mod_date_combo.addItem("自定义...", "custom")
        self.mod_date_combo.setToolTip(
            "按文件最后修改日期筛选：\n"
            "- 不限: 不限制修改日期\n"
            "- 今天: 只搜索今天修改过的文件\n"
            "- 最近7天: 搜索按钮7天内修改的文件\n"
            "- 最近30天: 30天内修改的文件\n"
            "- 自定义: 手动设置日期范围\n\n"
            "注意：此筛选基于文件的最后修改时间"
        )
        
        mod_date_layout.addWidget(mod_date_label)
        mod_date_layout.addWidget(self.mod_date_combo)
        date_size_layout.addLayout(mod_date_layout)
        
        file_size_layout = QVBoxLayout()
        file_size_label = QLabel("文件大小:")
        self.file_size_combo = QComboBox()
        self.file_size_combo.addItem("不限", (0, float('inf')))
        self.file_size_combo.addItem("小于 1MB", (0, 1024 * 1024))
        self.file_size_combo.addItem("1MB - 10MB", (1024 * 1024, 10 * 1024 * 1024))
        self.file_size_combo.addItem("大于 10MB", (10 * 1024 * 1024, float('inf')))
        self.file_size_combo.setToolTip(
            "按文件大小范围筛选：\n"
            "- 不限: 不限制文件大小\n"
            "- 小于 1MB: 功能筛选小于1MB的文件\n"
            "- 1MB - 10MB: 功能筛选1MB到10MB之间的文件\n"
            "- 大于 10MB: 大于10MB的文件\n\n"
            "注意：1MB = 1024KB = 1,048,576字节"
        )
        
        file_size_layout.addWidget(file_size_label)
        file_size_layout.addWidget(self.file_size_combo)
        date_size_layout.addLayout(file_size_layout)
        
        self.subfolders_check = QCheckBox("包含子文件夹")
        self.subfolders_check.setChecked(True)
        search_layout.addWidget(self.subfolders_check)
        
        search_layout.addLayout(date_size_layout)
        
        search_group.setLayout(search_layout)
        main_layout.addWidget(search_group)
        
        button_layout = QHBoxLayout()
        
        self.search_button = QPushButton("开始搜索")
        self.search_button.clicked.connect(self.start_search)
        
        self.cancel_button = QPushButton("取消搜索")
        self.cancel_button.setEnabled(False)
        self.cancel_button.clicked.connect(self.cancel_search_action)
        
        self.target_button = QPushButton("选择目标位置")
        self.target_button.clicked.connect(self.select_target_folder)
        
        self.copy_button = QPushButton("开始文件归集")
        self.copy_button.setEnabled(False)
        self.copy_button.clicked.connect(self.copy_files)
        
        self.delete_button = QPushButton("删除原文件")
        self.delete_button.setEnabled(False)
        self.delete_button.clicked.connect(self.delete_files)
        
        self.log_button = QPushButton("生成PDF日志")
        self.log_button.setEnabled(True)
        self.log_button.clicked.connect(self.generate_pdf_log)
        
        self.help_button = QPushButton("使用说明")
        self.help_button.clicked.connect(self.show_help)
        
        button_layout.addWidget(self.search_button)
        button_layout.addWidget(self.cancel_button)
        button_layout.addWidget(self.target_button)
        button_layout.addWidget(self.copy_button)
        button_layout.addWidget(self.delete_button)
        button_layout.addWidget(self.log_button)
        button_layout.addWidget(self.help_button)
        
        main_layout.addLayout(button_layout)
        
        results_group = QGroupBox("搜索结果")
        results_layout = QVBoxLayout()
        
        self.results_tree = QTreeWidget()
        self.results_tree.setHeaderLabels(["文件名", "路径", "大小", "修改日期"])
        self.results_tree.setColumnWidth(0, 300)
        self.results_tree.setColumnWidth(1, 400)
        self.results_tree.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.results_tree.itemDoubleClicked.connect(self.show_file_info)
        
        self.results_tree.setContextMenuPolicy(Qt.CustomContextMenu)
        self.results_tree.customContextMenuRequested.connect(self.show_context_menu)
        
        self.results_tree.header().setSectionResizeMode(0, QHeaderView.Interactive)
        self.results_tree.header().setSectionResizeMode(1, QHeaderView.Interactive)
        self.results_tree.header().setStretchLastSection(False)
        
        # 修改3: 设置搜索结果树最小高度为5行
        self.results_tree.setMinimumHeight(125)  # 大约5行高度
        
        results_layout.addWidget(self.results_tree)
        
        self.current_path_label = QLabel("当前搜索路径: ")
        self.current_path_label.setStyleSheet("color: #7f8c8d; font-style: italic;")
        results_layout.addWidget(self.current_path_label)
        
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        results_layout.addWidget(self.progress_bar)
        
        status_layout = QHBoxLayout()
        self.status_count_label = QLabel("已找到: 0 个文件")
        self.status_count_label.setStyleSheet("font-weight: bold; color: #2980b9;")
        status_layout.addWidget(self.status_count_label)
        status_layout.addStretch(1)
        
        results_layout.addLayout(status_layout)
        
        results_group.setLayout(results_layout)
        main_layout.addWidget(results_group, 1)  # 1表示这个部件可以扩展
        
        self.status_label = QLabel("就绪")
        self.status_label.setStyleSheet("color: #7f8c8d; font-size: 10pt;")
        main_layout.addWidget(self.status_label)
        
        self.central_widget.setLayout(main_layout)
    
    def add_log(self, action, file_path=None):
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.operation_log.append(f"{timestamp} - {action}")
        if file_path:
            self.operated_files.add(str(file_path))
    
    def remove_selected_folders(self):
        selected_items = self.folder_list.selectedItems()
        if not selected_items:
            return
            
        for item in selected_items:
            folder = item.text()
            if folder in self.search_folders:
                self.search_folders.remove(folder)
                self.add_log("删除选中的搜索文件夹", folder)
                
        self.update_folder_list()
    
    def cancel_search_action(self):
        self.cancel_search = True
        self.status_label.setText("搜索已取消")
        self.cancel_button.setEnabled(False)
        self.search_button.setEnabled(True)
        self.add_log("取消搜索")
    
    def add_search_folder(self):
        folder = QFileDialog.getExistingDirectory(self, "选择搜索文件夹")
        if folder and folder not in self.search_folders:
            self.search_folders.append(folder)
            self.update_folder_list()
            self.add_log(f"添加搜索文件夹: {folder}", folder)
    
    def add_drive(self):
        drives = [f"{d}:\\" for d in "ABCDEFGHIJKLMNOPQRSTUVWXYZ" if Path(f"{d}:").exists()]
        if not drives:
            QMessageBox.information(self, "信息", "未找到可用盘符")
            return
        
        menu = QMenu(self)
        for drive in drives:
            action = QAction(drive, self)
            action.triggered.connect(lambda checked, d=drive: self.add_drive_action(d))
            menu.addAction(action)
        
        pos = self.add_drive_button.mapToGlobal(QPoint(0, self.add_drive_button.height()))
        menu.exec_(pos)
    
    def add_drive_action(self, drive):
        if drive not in self.search_folders:
            self.search_folders.append(drive)
            self.update_folder_list()
            self.add_log(f"添加盘符: {drive}", drive)
    
    def clear_search_folders(self):
        self.search_folders = []
        self.update_folder_list()
        self.add_log("清空搜索文件夹列表")
    
    def update_folder_list(self):
        self.folder_list.clear()
        for folder in self.search_folders:
            self.folder_list.addItem(folder)
    
    def show_context_menu(self, position):
        menu = QMenu()
        
        open_file_action = QAction("打开文件", self)
        open_file_action.triggered.connect(self.open_selected_file)
        
        open_folder_action = QAction("打开所在文件夹", self)
        open_folder_action.triggered.connect(self.open_file_folder)
        
        menu.addAction(open_file_action)
        menu.addAction(open_folder_action)
        menu.exec_(self.results_tree.viewport().mapToGlobal(position))
    
    def open_selected_file(self):
        selected_items = self.results_tree.selectedItems()
        if not selected_items:
            return
        
        file_path = selected_items[0].data(0, Qt.UserRole)
        try:
            os.startfile(file_path)
            self.add_log(f"打开文件: {file_path}", file_path)
        except Exception as e:
            QMessageBox.warning(self, "错误", f"无法打开文件: {str(e)}")
    
    def open_file_folder(self):
        selected_items = self.results_tree.selectedItems()
        if not selected_items:
            return
        
        file_path = selected_items[0].data(0, Qt.UserRole)
        folder_path = Path(file_path).parent
        
        try:
            os.startfile(str(folder_path))
            self.add_log(f"打开所在文件夹: {folder_path}", folder_path)
        except Exception as e:
            QMessageBox.warning(self, "错误", f"无法打开文件夹: {str(e)}")
    
    def select_target_folder(self):
        folder = QFileDialog.getExistingDirectory(self, "选择目标文件夹")
        if folder:
            self.target_folder = folder
            self.status_label.setText(f"目标文件夹已设置为: {folder}")
            self.copy_button.setEnabled(True)
            self.add_log(f"设置目标文件夹: {folder}", folder)
    
    def matches_keyword(self, text, keyword):
        if not keyword:
            return True
        
        exact_matches = re.findall(r'"([^"]*)"', keyword)
        for exact in exact_matches:
            if exact.lower() not in text.lower():
                return False
            keyword = keyword.replace(f'"{exact}"', '')
        
        tokens = re.split(r'\s+', keyword.strip())
        must_include = []
        must_exclude = []
        any_include = []
        
        for token in tokens:
            if not token:
                continue
            if token.startswith('+'):
                must_include.append(token[1:].lower())
            elif token.startswith('-'):
                must_exclude.append(token[1:].lower())
            elif '|' in token:
                any_include.extend([t.lower() for t in token.split('|')])
            else:
                any_include.append(token.lower())
        
        for term in must_include:
            if term not in text.lower():
                return False
        
        for term in must_exclude:
            if term in text.lower():
                return False
        
        if any_include:
            found = False
            for term in any_include:
                if term in text.lower():
                    found = True
                    break
            if not found:
                return False
        
        return True
    
    # 修复1: 使用更可靠的文件占用检测方法
    def is_file_locked(self, filepath):
        """简化文件占用检测方法"""
        try:
            # 尝试以追加模式打开文件
            with open(filepath, 'ab') as f:
                return False
        except IOError:
            return True
        except Exception:  # 其他异常不视为占用
            return False
    
    def get_search_mode(self):
        if self.filename_radio.isChecked():
            return "filename"
        elif self.content_radio.isChecked():
            return "content"
        elif self.both_radio.isChecked():
            return "both"
        return "filename"
    
    # 新增内容搜索函数
    def search_content(self, file_path, keyword):
        try:
            file_path = Path(file_path)
            ext = file_path.suffix.lower()
            
            if ext in ['.txt', '.py', '.java', '.cpp', '.h', '.html', '.css', '.js', '.csv', '.ini', '.log']:
                return self.search_text_file(file_path, keyword)
            
            elif ext == '.pdf':
                return self.search_pdf(file_path, keyword)
            
            elif ext in ['.docx']:
                return self.search_docx(file_path, keyword)
            
            elif ext in ['.xlsx']:
                return self.search_excel(file_path, keyword)
            
            else:
                return False
                
        except Exception as e:
            print(f"内容搜索失败: {file_path} - {str(e)}")
            return False
    
    def search_text_file(self, file_path, keyword):
        encodings = ['utf-8', 'gbk', 'latin-1']
        for encoding in encodings:
            try:
                with file_path.open('r', encoding=encoding, errors='ignore') as f:
                    content = f.read(3000)
                    return self.matches_keyword(content, keyword)
            except UnicodeDecodeError:
                continue
        return False
    
    def search_pdf(self, file_path, keyword):
        try:
            from fitz import fitz  # 使用PyMuPDF
            doc = fitz.open(str(file_path))
            text = ""
            for page in doc:
                text += page.get_text()
                if len(text) > 3000:
                    break
            doc.close()
            return self.matches_keyword(text, keyword)
        except ImportError:
            return self.search_text_file(file_path, keyword)
        except Exception:
            return False
    
    def search_docx(self, file_path, keyword):
        try:
            from docx import Document
            doc = Document(str(file_path))
            text = ""
            for para in doc.paragraphs:
                text += para.text + " "
                if len(text) > 3000:
                    break
            return self.matches_keyword(text, keyword)
        except ImportError:
            return self.search_text_file(file_path, keyword)
        except Exception:
            return False
    
    def search_excel(self, file_path, keyword):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(str(file_path), read_only=True)
            text = ""
            for sheet in wb:
                for row in sheet.iter_rows(values_only=True):
                    for cell in row:
                        if cell:
                            text += str(cell) + " "
                    if len(text) > 3000:
                        break
                if len(text) > 3000:
                    break
            wb.close()
            return self.matches_keyword(text, keyword)
        except ImportError:
            return self.search_text_file(file_path, keyword)
        except Exception:
            return False
    
    def start_search(self):
        if not self.search_folders:
            QMessageBox.warning(self, "错误", "请添加至少一个搜索文件夹或盘符！")
            return
        
        self.search_button.setEnabled(False)
        self.cancel_button.setEnabled(True)
        self.cancel_search = False
        
        keyword = self.keyword_entry.text()
        file_types = self.filetype_combo.currentData()
        include_subfolders = self.subfolders_check.isChecked()
        search_mode = self.get_search_mode()
        
        self.add_log(f"开始搜索，关键词: {keyword}，文件类型: {self.filetype_combo.currentText()}")
        
        if file_types == "custom":
            custom_types, ok = QInputDialog.getText(self, "自定义文件类型", 
                                                   "请输入文件扩展名，用分号分隔：\n例如: .py;.java;.cpp", 
                                                   text="")
            if not ok or not custom_types:
                self.search_button.setEnabled(True)
                self.cancel_button.setEnabled(False)
                return
            file_types = [ext.strip().lower() for ext in custom_types.split(';') if ext.strip()]
        
        self.search_results = []
        self.results_tree.clear()
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)
        self.found_files_count = 0
        self.status_count_label.setText("已找到: 0 个文件")
        self.status_label.setText("正在搜索文件...")
        QApplication.processEvents()
        
        size_range = self.file_size_combo.currentData()
        mod_date_range = self.mod_date_combo.currentData()
        
        # 单次扫描方案 - 移除预扫描计数
        processed = 0
        total_files = 0  # 不再使用总文件数计算进度
        
        for folder in self.search_folders:
            if self.cancel_search:
                break
                
            folder_path = Path(folder)
            if not folder_path.exists():
                continue
                
            if include_subfolders:
                walk_iter = os.walk(folder)
            else:
                try:
                    files = [f for f in folder_path.iterdir() if f.is_file()]
                    walk_iter = [(str(folder_path), [], [f.name for f in files])]
                except Exception as e:
                    print(f"访问文件夹出错: {folder} - {str(e)}")
                    continue
            
            for root, _, files in walk_iter:
                if self.cancel_search:
                    break
                    
                # 更新当前搜索路径显示
                self.current_path_label.setText(f"当前搜索路径: {root}")
                self.status_label.setText(f"正在搜索: {root}")
                QApplication.processEvents()
                
                for file in files:
                    if self.cancel_search:
                        break
                        
                    file_path = Path(root) / file
                    
                    try:
                        if not file_path.exists() or not os.access(str(file_path), os.R_OK):
                            continue
                            
                        file_stat = file_path.stat()
                        file_size = file_stat.st_size
                        mod_date = datetime.datetime.fromtimestamp(file_stat.st_mtime).date()
                        
                        if not (size_range[0] <= file_size <= size_range[1]):
                            continue
                        
                        if mod_date_range != (None, None) and mod_date_range != "custom":
                            start_date, end_date = mod_date_range
                            if not (start_date <= mod_date <= end_date):
                                continue
                        
                        ext = file_path.suffix.lower()
                        if file_types and ext not in file_types:
                            continue
                        
                        filename_match = False
                        content_match = False
                        
                        if keyword:
                            if search_mode in ["filename", "both"]:
                                filename_match = self.matches_keyword(file, keyword)
                            
                            if search_mode in ["content", "both"] and not filename_match:
                                content_match = self.search_content(str(file_path), keyword)
                        
                        if keyword:
                            if search_mode == "filename":
                                if not filename_match:
                                    continue
                            elif search_mode == "content":
                                if not content_match:
                                    continue
                            elif search_mode == "both":
                                if not (filename_match or content_match):
                                    continue
                        
                        self.search_results.append({
                            'path': str(file_path),
                            'name': file,
                            'size': file_size,
                            'mod_date': mod_date.strftime("%Y-%m-%d")
                        })
                        
                        item = QTreeWidgetItem([
                            file, 
                            root, 
                            self.format_size(file_size), 
                            mod_date.strftime("%Y-%m-%d")
                        ])
                        item.setData(0, Qt.UserRole, str(file_path))
                        item.setToolTip(0, file)
                        item.setToolTip(1, root)
                        self.results_tree.addTopLevelItem(item)
                        
                        self.found_files_count += 1
                        # 实时更新找到的文件数量
                        self.status_count_label.setText(f"已找到: {self.found_files_count} 个文件")
                        QApplication.processEvents()
                        
                    except Exception as e:
                        print(f"跳过文件 {file_path}，原因: {str(e)}")
                        continue
                    
                    processed += 1
                    if processed % 100 == 0:
                        print(f"已处理 {processed} 个文件，当前路径: {root}")  # 调试日志
                        QApplication.processEvents()  # 仅保持UI响应
                        if self.cancel_search:
                            break
        
        self.progress_bar.setVisible(False)  # 保持进度条隐藏
        self.current_path_label.setText("当前搜索路径: ")
        
        if self.cancel_search:
            self.status_label.setText(f"搜索已取消，已找到 {self.found_files_count} 个文件")
            self.add_log("搜索已取消")
        else:
            self.status_label.setText(f"搜索完成，找到 {self.found_files_count} 个文件")
            self.add_log(f"搜索完成，找到 {self.found_files_count} 个文件")
        
        self.status_count_label.setText(f"已找到: {self.found_files_count} 个文件")
        self.copy_button.setEnabled(bool(self.search_results))
        self.search_button.setEnabled(True)
        self.cancel_button.setEnabled(False)
    
    def format_size(self, size):
        if size < 1024:
            return f"{size} B"
        elif size < 1024 * 1024:
            return f"{size/1024:.2f} KB"
        elif size < 1024 * 1024 * 1024:
            return f"{size/(1024*1024):.2f} MB"
        else:
            return f"{size/(1024*1024*1024):.2f} GB"
    
    def copy_files(self):
        if not self.target_folder:
            QMessageBox.warning(self, "错误", "请先选择目标文件夹！")
            return
        
        if not self.search_results:
            QMessageBox.warning(self, "错误", "没有可复制的文件！")
            return
        
        # 检查目标文件夹是否有写入权限
        target_path = Path(self.target_folder)
        if not target_path.exists() or not os.access(str(target_path), os.W_OK):
            QMessageBox.critical(self, "错误", "目标文件夹无写入权限！")
            return
        
        target_path.mkdir(parents=True, exist_ok=True)
        
        existing_files = set()
        for file_info in self.search_results:
            target_file = target_path / file_info['name']
            if target_file.exists():
                existing_files.add(file_info['name'])
        
        if not existing_files:
            self.copy_files_without_conflicts()
            return
        
        conflict_dialog = FileConflictDialog(self, self.search_results, self.target_folder)
        if conflict_dialog.exec_() == QDialog.Accepted:
            files_to_copy = conflict_dialog.get_selected_files()
            self.copy_selected_files(files_to_copy)
    
    def copy_files_without_conflicts(self):
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)
        self.status_label.setText("正在复制文件...")
        QApplication.processEvents()
        
        error_files = []
        
        for i, file_info in enumerate(self.search_results):
            src = Path(file_info['path'])
            dst = Path(self.target_folder) / file_info['name']
            
            try:
                # 修复2: 直接尝试复制
                shutil.copy2(str(src), str(dst))
                self.add_log(f"复制文件到目标文件夹: {dst}", src)
            except Exception as e:
                error_files.append(f"{src} ({str(e)})")
            
            progress = int((i + 1) / len(self.search_results) * 100)
            self.progress_bar.setValue(progress)
            if i % 10 == 0:
                QApplication.processEvents()
        
        self.progress_bar.setVisible(False)
        
        if error_files:
            error_msg = "以下文件复制失败：\n\n" + "\n".join(error_files[:10])
            if len(error_files) > 10:
                error_msg += f"\n\n...以及另外 {len(error_files)-10} 个文件"
            QMessageBox.warning(self, "复制错误", error_msg)
            self.status_label.setText(f"已成功复制 {len(self.search_results)-len(error_files)} 个文件，{len(error_files)} 个失败")
        else:
            self.status_label.setText(f"已成功复制 {len(self.search_results)} 个文件到目标文件夹")
            self.delete_button.setEnabled(True)
    
    def copy_selected_files(self, files_to_copy):
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)
        self.status_label.setText("正在复制文件...")
        QApplication.processEvents()
        
        error_files = []
        
        for i, file_info in enumerate(files_to_copy):
            src = Path(file_info['path'])
            dst = Path(self.target_folder) / file_info['new_name']
            
            try:
                # 修复2: 直接尝试复制
                shutil.copy2(str(src), str(dst))
                self.add_log(f"复制文件到目标文件夹: {dst}", src)
            except Exception as e:
                error_files.append(f"{src} ({str(e)})")
            
            progress = int((i + 1) / len(files_to_copy) * 100)
            self.progress_bar.setValue(progress)
            if i % 10 == 0:
                QApplication.processEvents()
        
        self.progress_bar.setVisible(False)
        
        if error_files:
            error_msg = "以下文件复制失败：\n\n" + "\n".join(error_files[:10])
            if len(error_files) > 10:
                error_msg += f"\n\n...以及另外 {len(error_files)-10} 个文件"
            QMessageBox.warning(self, "复制错误", error_msg)
            self.status_label.setText(f"已成功复制 {len(files_to_copy)-len(error_files)} 个文件，{len(error_files)} 个失败")
        else:
            self.status_label.setText(f"已成功复制 {len(files_to_copy)} 个文件到目标文件夹")
            self.delete_button.setEnabled(True)
    
    def delete_files(self):
        if not self.search_results:
            QMessageBox.warning(self, "错误", "没有可删除的文件！")
            return
        
        # 添加警告弹窗和免责声明
        disclaimer = (
            "您即将删除原文件，请谨慎操作\n\n"
            "【免责声明】本软件为免费工具，用户自愿使用。开发者不承诺软件绝对安全，"
            "对因使用软件导致的数据丢失、系统损坏等后果不承担责任。"
            "禁止将软件用于非法目的。"
        )
        reply = QMessageBox.warning(self, "警告", disclaimer,
                                  QMessageBox.Ok | QMessageBox.Cancel)
        
        if reply != QMessageBox.Ok:
            return
            
        # 添加确认对话框
        reply = QMessageBox.question(self, "确认删除", 
                                    "确定要永久删除原文件吗？此操作不可撤销！",
                                    QMessageBox.Yes | QMessageBox.No)
        
        if reply == QMessageBox.Yes:
            self.progress_bar.setVisible(True)
            self.progress_bar.setValue(0)
            self.status_label.setText("正在删除文件...")
            QApplication.processEvents()
            
            success_count = 0
            error_files = []
            total_files = len(self.search_results)
            update_interval = max(1, total_files // 50)
            
            # 创建临时列表存储删除结果
            remaining_files = []
            
            for i, file_info in enumerate(self.search_results):
                file_path = Path(file_info['path'])
                try:
                    if not file_path.exists():
                        error_files.append(f"{file_path} (文件不存在)")
                        remaining_files.append(file_info)
                        continue
                        
                    # 修复3: 使用os.remove直接删除
                    file_path.unlink()
                    self.add_log(f"删除文件: {file_path}", file_path)
                    success_count += 1
                    
                    # 从树控件中移除
                    for index in range(self.results_tree.topLevelItemCount()):
                        item = self.results_tree.topLevelItem(index)
                        if item.data(0, Qt.UserRole) == str(file_path):
                            self.results_tree.takeTopLevelItem(index)
                            break
                except Exception as e:
                    error_files.append(f"{file_path} ({str(e)})")
                    remaining_files.append(file_info)
                
                if i % update_interval == 0 or i == total_files - 1:
                    progress = int((i + 1) / total_files * 100)
                    self.progress_bar.setValue(progress)
                    QApplication.processEvents()
            
            # 更新搜索结果列表
            self.search_results = remaining_files
            
            self.progress_bar.setVisible(False)
            
            if error_files:
                error_msg = "以下文件删除失败：\n\n" + "\n".join(error_files[:10])
                if len(error_files) > 10:
                    error_msg += f"\n\n...以及另外 {len(error_files)-10} 个文件"
                QMessageBox.warning(self, "删除错误", 
                                   f"{error_msg}\n\n请手动删除这些文件。")
            
            self.status_label.setText(f"已删除 {success_count}/{total_files} 个文件")
            self.delete_button.setEnabled(bool(self.search_results))
    
    def _wrap_text(self, text, max_len=40):
        """将长文本自动换行，max_len为每行最大字符数"""
        if not text or len(text) <= max_len:
            return text
        lines = [text[i:i+max_len] for i in range(0, len(text), max_len)]
        return "<br/>".join(lines)
    
    def _extract_filename_for_log(self, log):
        """从日志字符串中提取文件名（不含路径），如无则原样返回"""
        match = re.search(r'([A-Z]:\\[^\s]+)', log)
        if match:
            path = match.group(1)
            filename = Path(path).name
            return log.replace(path, filename)
        return log

    def generate_pdf_log(self):
        options = ["前20条", "前50条", "全部"]
        choice, ok = QInputDialog.getItem(self, "导出操作记录", "请选择导出操作过程的条数：", options, 0, False)
        if not ok:
            return
        if choice == "前20条":
            op_logs = self.operation_log[:20]
        elif choice == "前50条":
            op_logs = self.operation_log[:50]
        else:
            op_logs = self.operation_log
            if len(self.operation_log) > 100:
                QMessageBox.warning(self, "提示", "导出全部操作记录可能导致PDF文档过大，请耐心等待。")

        file_path, _ = QFileDialog.getSaveFileName(self, "保存PDF日志", "", "PDF文件 (*.pdf)")
        if not file_path:
            return

        file_path = Path(file_path)
        if file_path.suffix.lower() != ".pdf":
            file_path = file_path.with_suffix(".pdf")

        try:
            doc = SimpleDocTemplate(str(file_path), pagesize=letter)
            styles = getSampleStyleSheet()
            story = []

            title_style = styles['Title']
            title_style.fontName = 'SimSun'

            heading_style = styles['Heading2']
            heading_style.fontName = 'SimSun'

            normal_style = styles['Normal']
            normal_style.fontName = 'SimSun'

            title = Paragraph(f"<b>文件归集管理器操作日志</b>", title_style)
            story.append(title)
            story.append(Spacer(1, 12))

            search_mode = self.get_search_mode()
            search_mode_text = {
                "filename": "仅文件名",
                "content": "仅内容",
                "both": "两者同时"
            }.get(search_mode, "仅文件名")

            info_data = [
                ["操作日期", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
                ["软件版本", f"文件归集管理器 V{self.version}"],
                ["搜索文件夹", "\n".join(self.search_folders)],
                ["目标文件夹", self.target_folder or "未设置"],
                ["关键词", self.keyword_entry.text() or "无"],
                ["搜索模式", search_mode_text],
                ["文件类型", self.filetype_combo.currentText()],
                ["文件数量", str(len(self.search_results))]
            ]
            info_table = Table(info_data, colWidths=[150, 350])
            info_table.setStyle(TableStyle([
                ('FONT', (0, 0), (-1, -1), 'SimSun', 10),
                ('BACKGROUND', (0, 0), (0, -1), colors.lightblue),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey)
            ]))
            story.append(info_table)
            story.append(Spacer(1, 24))

            # 操作过程记录表格（加序号，自动换行）
            if op_logs:
                op_title = Paragraph("<b>操作过程记录</b>", heading_style)
                story.append(op_title)
                story.append(Spacer(1, 12))
                op_data = [["序号", "时间与操作"]]
                for idx, log in enumerate(op_logs, 1):
                    log_short = self._extract_filename_for_log(log)
                    op_data.append([
                        str(idx),
                        Paragraph(log_short, normal_style)
                    ])
                op_table = Table(op_data, colWidths=[40, 460])
                op_table.setStyle(TableStyle([
                    ('FONT', (0, 0), (-1, 0), 'SimSun', 10),
                    ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                    ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                    ('FONT', (0, 1), (-1, -1), 'SimSun', 9),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP')
                ]))
                story.append(op_table)
                story.append(Spacer(1, 24))

            # 操作过的文件名表格（加序号，自动换行）
            if self.operated_files:
                file_title = Paragraph("<b>操作过的文件名</b>", heading_style)
                story.append(file_title)
                story.append(Spacer(1, 12))
                file_data = [["序号", "文件名"]]
                for idx, fname in enumerate(sorted(self.operated_files), 1):
                    file_data.append([
                        str(idx),
                        Paragraph(fname, normal_style)
                    ])
                file_table = Table(file_data, colWidths=[40, 460])
                file_table.setStyle(TableStyle([
                    ('FONT', (0, 0), (-1, 0), 'SimSun', 10),
                    ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                    ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                    ('FONT', (0, 1), (-1, -1), 'SimSun', 9),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP')
                ]))
                story.append(file_table)
                story.append(Spacer(1, 24))

            # 文件列表
            if self.search_results:
                files_title = Paragraph("<b>文件列表</b>", heading_style)
                story.append(files_title)
                story.append(Spacer(1, 12))

                file_data = [["文件名", "大小", "修改日期"]]
                for file_info in self.search_results:
                    file_data.append([
                        Paragraph(file_info['name'], normal_style),
                        self.format_size(file_info['size']),
                        file_info['mod_date']
                    ])

                file_table = Table(file_data, colWidths=[350, 100, 100])
                file_table.setStyle(TableStyle([
                    ('FONT', (0, 0), (-1, 0), 'SimSun', 10),
                    ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                    ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                    ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
                    ('ALIGN', (2, 0), (2, -1), 'CENTER'),
                    ('FONT', (0, 1), (-1, -1), 'SimSun', 9),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP')
                ]))
                story.append(file_table)
                story.append(Spacer(1, 24))

            # 版本更新记录
            updates_title = Paragraph("<b>版本更新记录</b>", heading_style)
            story.append(updates_title)
            story.append(Spacer(1, 12))

            updates_data = [
                ["版本", "日期", "更新内容"],
                ["V2.3.4", "2025-07-17", "优化搜索结果及按钮显示"],
                ["V2.3.3", "2025-07-10", "更改按钮文字，优化显示效果——感谢薯友\"疯\"提供的意见"],
                ["V2.3.2", "2025-07-07", "将应用程序图标从绘制改为使用win.ico文件"],
                ["V2.3.1", "2025-07-06", "优化PDF日志导出，操作过程日志仅显示文件名并自动换行。"],
                ["V2.3", "2025-07-05", "允许用户垂直调整窗口高度，改进文件列表的显示优化"],
                ["V2.2", "2025-07-04", "添加应用程序图标，优化界面显示"],
                ["V2.1", "2025-07-03", "修复内容搜索功能，改进PDF处理，增强文件占用检测"],
                ["V2.0", "2025-07-03", "修复文件复制功能，优化性能"],
                ["V1.3.1", "2025-07-03", "修复文件占用检测，优化性能，解决界面卡顿问题"],
                ["V1.3", "2025-07-03", "增加搜索模式选择（文件名/内容/两者），优化搜索性能，增加当前路径显示"],
                ["V1.2", "2025-07-03", "修复搜索功能核心问题，优化用户界面，改进文件处理逻辑"],
                ["V1.1", "2025-07-02", "新增多文件夹/盘符搜索、高级搜索语法、多语言PDF支持、右键菜单等功能"],
                ["V1.0", "2025-07-02", "初始版本发布，包含基本搜索和文件管理功能"]
            ]
            updates_table = Table(updates_data, colWidths=[80, 100, 320])
            updates_table.setStyle(TableStyle([
                ('FONT', (0, 0), (-1, 0), 'SimSun', 10),
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ('FONT', (0, 1), (-1, -1), 'SimSun', 9)
            ]))
            story.append(updates_table)

            doc.build(story)
            self.status_label.setText(f"PDF日志已保存到: {file_path}")
            self.add_log(f"生成PDF日志: {file_path}")
            QMessageBox.information(self, "成功", "PDF日志已成功生成！")

        except Exception as e:
            QMessageBox.critical(self, "错误", f"生成PDF日志时出错: {str(e)}")
            print(traceback.format_exc())
    
    def show_file_info(self, item, column):
        file_path = item.data(0, Qt.UserRole)
        if not file_path:
            return
        
        file_path = Path(file_path)
        try:
            file_stat = file_path.stat()
            size = self.format_size(file_stat.st_size)
            mod_date = datetime.datetime.fromtimestamp(file_stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S")
            create_date = datetime.datetime.fromtimestamp(file_stat.st_ctime).strftime("%Y-%m-%d %H:%M:%S")
            
            owner = "未知"
            try:
                import win32security
                sd = win32security.GetFileSecurity(str(file_path), win32security.OWNER_SECURITY_INFORMATION)
                owner_sid = sd.GetSecurityDescriptorOwner()
                owner, _, _ = win32security.LookupAccountSid(None, owner_sid)
            except:
                pass
            
            content_preview = ""
            try:
                encodings = ['utf-8', 'gbk', 'latin-1']
                for encoding in encodings:
                    try:
                        with file_path.open('r', encoding=encoding, errors='ignore') as f:
                            content_preview = f.read(300)
                            break
                    except UnicodeDecodeError:
                        continue
            except:
                content_preview = "无法预览文件内容"
            
            file_status = "状态: 可用"
            if self.is_file_locked(str(file_path)):
                file_status = "状态: <font color='red'>被其他程序占用</font>"
            
            info = f"""
            <b>文件信息</b>
            <table>
            <tr><td><b>文件名：</b></td><td>{file_path.name}</td></tr>
            <tr><td><b>路径：</b></td><td>{file_path}</td></tr>
            <tr><td><b>大小：</b></td><td>{size}</td></tr>
            <tr><td><b>修改日期：</b></td><td>{mod_date}</td></tr>
            <tr><td><b>创建日期：</b></td><td>{create_date}</td></tr>
            <tr><td><b>所有者：</b></td><td>{owner}</td></tr>
            <tr><td><b>{file_status}</b></td><td></td></tr>
            </table>
            <b>内容预览 (前300字符)：</b>
            <pre>{content_preview}</pre>
            """
            
            QMessageBox.information(self, "文件详情", info)
            self.add_log(f"查看文件信息: {file_path}", file_path)
        except Exception as e:
            QMessageBox.warning(self, "错误", f"无法获取文件信息: {str(e)}")
    
    def show_help(self):
        help_text = f"""
        <div style="color: red; font-weight: bold; border: 1px solid red; padding: 10px; margin-bottom: 20px;">
        【免责声明】本软件为免费工具，用户自愿使用。开发者不承诺软件绝对安全，
        对因使用软件导致的数据丢失、系统损坏等后果不承担责任。
        禁止将软件用于非法目的。
        </div>
        
        <h2>文件归集管理器 V{self.version} 使用说明</h2>
        
        <h3>基本功能</h3>
        <p>文件归集管理器是一个强大的文件集中管理工具，可以帮助您搜索、整理和管理计算机中的文件。</p>
        
        <h3>操作步骤</h3>
        <ol>
            <li><b>添加搜索文件夹/盘符</b>：使用"添加文件夹"或"添加盘符"按钮选择搜索范围</li>
            <li><b>设置搜索条件</b>： 
                <ul>
                    <li>关键词：支持高级搜索语法（将鼠标悬停在输入框上查看详情）</li>
                    <li>搜索模式：选择文件名/内容/两者同时搜索</li>
                    <li>文件类型：按类别筛选文件（将鼠标悬停在选择框上查看详情）</li>
                    <li>修改日期：按时间范围筛选文件（将鼠标悬停在选择框上查看详情）</li>
                    <li>文件大小：按大小范围筛选文件（将鼠标悬停在选择框上查看详情）</li>
                </ul>
            </li>
            <li><b>开始搜索</b>：点击"开始搜索"按钮执行搜索</li>
            <li><b>选择目标文件夹</b>：点击"选择目标文件夹"按钮指定文件复制位置</li>
            <li><b>复制文件</b>：点击"复制文件到目标文件夹"按钮复制文件</li>
            <li><b>删除原文件：可选步骤，永久删除原文件（不可撤销）</b></li>
            <li><b>生成PDF日志</b>：创建操作记录PDF文件</li>
        </ol>
        
        <h3>高级功能</h3>
        <ul>
            <li><b>搜索模式选择</b>：灵活选择仅文件名/仅内容/两者同时搜索</li>
            <li><b>多文件夹搜索</b>：支持同时搜索多个文件夹或整个盘符</li>
            <li><b>右键菜单</b>：在搜索结果上右键可打开文件或所在文件夹</li>
            <li><b>冲突处理</b>：自动检测目标文件夹中的同名文件并提供解决方案</li>
            <li><b>多语言支持</b>：PDF日志支持中文、英文、日文等多种语言</li>
            <li><b>文件占用检测</b>：自动检测并提示被占用的文件</li>
            <li><b>搜索状态显示</b>：实时显示当前搜索的文件夹路径</li>
            <li><b>美观界面</b>：简洁现代的UI设计，包含应用程序图标，可垂直改变窗口大小</li>
        </ul>
        
        <h3>文件冲突处理</h3>
        <p>当目标文件夹中存在同名文件时，系统会提示您选择处理方式：</p>
        <ul>
            <li><b>覆盖</b>：替换目标文件夹中的文件</li>
            <li><b>跳过</b>：保留目标文件夹中的文件</li>
            <li><b>重命名</b>：为新文件添加后缀避免冲突</li>
        </ul>
        
        <h3>注意事项</h3>
        <ul>
            <li>删除操作会永久删除文件，不可撤销</li>
            <li>内容搜索仅检查文件前3000个字符</li>
            <li>本软件仅支持Windows系统</li>
            <li>对于无法访问的文件会自动跳过</li>
        </ul>

        <h3>内容搜索支持的文件格式</h3>
        <ul>
            <li><b>文本文件</b>：.txt, .py, .java, .cpp, .h, .html, .css, .js, .csv, .ini, .log</li>
            <li><b>PDF文档</b>：.pdf（需要安装PyMuPDF库）</li>
            <li><b>Word文档</b>：.docx（需要安装python-docx库）</li>
            <li><b>Excel表格</b>：.xlsx（需要安装openpyxl库）</li>
            <li><b>其他文件</b>：跳过内容搜索</li>
        </ul>
        
        <h3>版本更新记录</h3>
        <p><b>v2.3.4(2025-07-17)</b></p>
        <ul>
            <li>优化搜索结果及按钮显示</li>
        </ul>
        <p><b>v2.3.3(2025-07-10)</b></p>
        <ul>
            <li>更改按钮文字，优化显示效果——感谢薯友"疯"提供的意见</li>
        </ul>
        <p><b>v2.3.2(2025-07-07)</b></p>
        <ul>
            <li>更新软件图标，将应用程序图标从绘制改为使用ico文件。<li>
        </ul>
        <p><b>V2.3.1 (2025-07-06)</b></p>
        <ul>
            <li>优化PDF日志导出，操作过程以表格形式展示，去除无匹配文件提示，完善帮助文档。<li>
            <li>导出PDF日志时可选择导出操作记录条数，支持导出所有操作过的文件名，优化大日志导出体验。</li>
        </ul>
        <p><b>V2.3 (2025-07-05)</b></p>
        <ul>
            <li>允许用户垂直调整窗口高度，改进文件列表的显示优化</li>
            <li>（特别感谢用户QR的宝贵建议\u2764）</li>
        </ul>
        <p><b>V2.2 (2025-07-04)</b></p>
        <ul>
            <li>添加应用程序图标，优化界面显示</li>
            <li>修复文件删除功能，改进错误提示</li>
        </ul>
        
        <p><b>V2.1 (2025-07-03)</b></p>
        <ul>
            <li>修复内容搜索功能</li>
            <li>改进PDF处理</li>
            <li>增强文件占用检测</li>
        </ul>
        
        <p><b>V2.0 (2025-07-03)</b></p>
        <ul>
            <li>修复文件复制功能</li>
            <li>优化性能</li>
        </ul>
        
        <p><b>V1.3.1 (2025-07-03)</b></p>
        <ul>
            <li>修复文件占用检测问题</li>
            <li>优化性能，解决界面卡顿</li>
            <li>改进文件复制和删除的错误处理</li>
            <li>固定窗口尺寸，优化长文件名显示</li>
        </ul>
        
        <p><b>V1.3 (2025-07-03)</b></p>
        <ul>
            <li>增加搜索模式选择：仅文件名/仅内容/两者同时</li>
            <li>内容搜索扩展至前300字符</li>
            <li>优化搜索逻辑，提高准确性</li>
            <li>在PDF日志中增加搜索模式信息</li>
            <li>改进帮助文档和工具提示</li>
        </ul>
        
        <p><b>V1.2 (2025-07-03)</b></p>
        <ul>
            <li>修复搜索功能核心问题</li>
            <li>优化用户界面和操作体验</li>
            <li>改进文件处理逻辑</li>
            <li>更新软件名称为"文件归集管理器"</li>
            <li>更新版权信息为"D&Ai/2FX 版权所有"</li>
            <li>修复PDF日志生成问题</li>
            <li>增强内容搜索功能</li>
            <li>解决文件删除参数错误问题</li>
        </ul>
        
        <p><b>V1.1 (2025-07-02)</b></p>
        <ul>
            <li>新增多文件夹/盘符搜索功能</li>
            <li>添加高级搜索语法支持</li>
            <li>改进PDF多语言支持</li>
            <li>增加右键菜单功能</li>
            <li>优化文件冲突检测逻辑</li>
            <li>添加更多提示和帮助信息</li>
        </ul>
        
        <p><b>V1.0 (2025-07-02)</b></p>
        <ul>
            <li>初始版本制作</li>
            <li>基本文件搜索和管理功能</li>
            <li>PDF日志生成</li>
            <li>文件冲突处理</li>
        </ul>
        
       
        """
        
        help_dialog = QDialog(self)
        help_dialog.setWindowTitle("使用说明")
        help_dialog.setGeometry(200, 200, 700, 600)
        
        layout = QVBoxLayout()
        
        text_edit = QLabel(help_text)
        text_edit.setWordWrap(True)
        
        scroll_area = QScrollArea()
        scroll_area.setWidget(text_edit)
        scroll_area.setWidgetResizable(True)
        
        layout.addWidget(scroll_area)
        
        close_button = QPushButton("关闭")
        close_button.clicked.connect(help_dialog.accept)
        layout.addWidget(close_button, alignment=Qt.AlignCenter)
        
        help_dialog.setLayout(layout)
        help_dialog.exec_()
    
class FileConflictDialog(QDialog):
    def __init__(self, parent, files, target_folder):
        super().__init__(parent)
        self.setWindowTitle("处理文件冲突")
        self.setGeometry(300, 300, 800, 500)
        
        self.files = files
        self.target_folder = target_folder
        self.selected_files = []
        
        layout = QVBoxLayout()
        
        title = QLabel("<b>检测到目标文件夹中存在同名文件，请选择处理方式：</b>")
        layout.addWidget(title)
        
        self.file_list = QListWidget()
        self.file_list.setSelectionMode(QAbstractItemView.ExtendedSelection)
        
        for file_info in files:
            file_name = file_info['name']
            target_path = Path(target_folder) / file_name
            
            conflict = "存在冲突" if target_path.exists() else ""
            
            item = QListWidgetItem(f"{file_name} - {conflict}")
            item.setData(Qt.UserRole, file_info)
            self.file_list.addItem(item)
            
            if conflict:
                item.setForeground(QColor("red"))
        
        layout.addWidget(self.file_list)
        
        help_label = QLabel("提示: 选择文件后点击下方按钮设置处理方式")
        help_label.setStyleSheet("color: #7f8c8d; font-style: italic;")
        layout.addWidget(help_label)
        
        button_layout = QHBoxLayout()
        
        self.overwrite_button = QPushButton("覆盖")
        self.overwrite_button.setToolTip("替换目标文件夹中的文件")
        self.overwrite_button.clicked.connect(lambda: self.set_action("overwrite"))
        
        self.skip_button = QPushButton("跳过")
        self.skip_button.setToolTip("不复制此文件")
        self.skip_button.clicked.connect(lambda: self.set_action("skip"))
        
        self.rename_button = QPushButton("重命名")
        self.rename_button.setToolTip("复制文件并重命名")
        self.rename_button.clicked.connect(lambda: self.set_action("rename"))
        
        self.auto_rename_button = QPushButton("全部重命名")
        self.auto_rename_button.setToolTip("为所有冲突文件添加后缀")
        self.auto_rename_button.clicked.connect(self.auto_rename_all)
        
        self.overwrite_all_button = QPushButton("全部覆盖")
        self.overwrite_all_button.setToolTip("将所有冲突文件全部覆盖")
        self.overwrite_all_button.clicked.connect(self.overwrite_all)
        
        button_layout.addWidget(self.overwrite_button)
        button_layout.addWidget(self.skip_button)
        button_layout.addWidget(self.rename_button)
        button_layout.addWidget(self.auto_rename_button)
        button_layout.addWidget(self.overwrite_all_button)
        
        layout.addLayout(button_layout)
        
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        
        layout.addWidget(button_box)
        
        self.setLayout(layout)
    
    def set_action(self, action):
        selected_items = self.file_list.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "警告", "请先选择文件！")
            return

        for item in selected_items:
            file_info = item.data(Qt.UserRole)

            if action == "overwrite":
                item.setText(f"{file_info['name']} - 将覆盖")
                item.setForeground(QColor("blue"))
                file_info['action'] = "overwrite"
                file_info['new_name'] = file_info['name']
            elif action == "skip":
                item.setText(f"{file_info['name']} - 将跳过")
                item.setForeground(QColor("gray"))
                file_info['action'] = "skip"
            elif action == "rename":
                file_path = Path(file_info['name'])
                base = file_path.stem
                ext = file_path.suffix
                counter = 1
                new_name = f"{base}_{counter}{ext}"
                target_path = Path(self.target_folder) / new_name
                while target_path.exists():
                    counter += 1
                    new_name = f"{base}_{counter}{ext}"
                    target_path = Path(self.target_folder) / new_name
                item.setText(f"{file_info['name']} -> {new_name}")
                file_info['action'] = "rename"
                file_info['new_name'] = new_name
                item.setForeground(QColor("darkgreen"))
    
    def auto_rename_all(self):
        for index in range(self.file_list.count()):
            item = self.file_list.item(index)
            file_info = item.data(Qt.UserRole)
            
            if "存在冲突" in item.text():
                file_path = Path(file_info['name'])
                base = file_path.stem
                ext = file_path.suffix
                counter = 1
                new_name = f"{base}_{counter}{ext}"
                
                # 确保新文件名唯一
                target_path = Path(self.target_folder) / new_name
                while target_path.exists():
                    counter += 1
                    new_name = f"{base}_{counter}{ext}"
                    target_path = Path(self.target_folder) / new_name
                
                item.setText(f"{file_info['name']} -> {new_name}")
                file_info['action'] = "rename"
                file_info['new_name'] = new_name
                item.setForeground(QColor("darkgreen"))
    
    def overwrite_all(self):
        for index in range(self.file_list.count()):
            item = self.file_list.item(index)
            file_info = item.data(Qt.UserRole)
            item.setText(f"{file_info['name']} - 将覆盖")
            item.setForeground(QColor("blue"))
            file_info['action'] = "overwrite"
            file_info['new_name'] = file_info['name']
    
    def get_selected_files(self):
        result = []
        for file_info in self.files:
            action = file_info.get('action', 'overwrite')
            if action == "skip":
                continue
            if action == "rename":
                if 'new_name' not in file_info:
                    file_path = Path(file_info['name'])
                    base = file_path.stem
                    ext = file_path.suffix
                    counter = 1
                    new_name = f"{base}_{counter}{ext}"
                    target_path = Path(self.target_folder) / new_name
                    while target_path.exists():
                        counter += 1
                        new_name = f"{base}_{counter}{ext}"
                        target_path = Path(self.target_folder) / new_name
                    file_info['new_name'] = new_name
            else:
                file_info['new_name'] = file_info['name']
            result.append(file_info)
        return result


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = FileGatherPro()
    window.show()
    sys.exit(app.exec_())
